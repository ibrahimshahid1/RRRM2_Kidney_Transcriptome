"""Repair A -- kinome-wide KSEA from the Johnson 2023 Ser/Thr atlas.

Replaces the three-substrate curated WNK--SPAK/OSR1 net (which can only ever
*confirm* the gene that was hand-curated into it) with a motif-driven assignment
over the *entire* OSD-462 phosphoproteome, then runs the existing Casado-style
KSEA (:func:`src.multiomics.regulator_activity.ksea`). The question becomes
unbiased: of all 303 Ser/Thr kinases scoreable from substrate motifs, which come
back inferred-*down* in flight -- and do SPAK (atlas label ``STLK3``) and OSR1
fall out near the top without having been planted there?

Pipeline
--------
1. ``load_atlas_pssms`` -- read the 303 position-normalised, log2-scaled PSSMs
   (sheet ``ser_thr_all_norm_scaled_matrice``) into a (kinase x position x
   residue) tensor of log2 weights.
2. ``score_sites`` -- for each phosphosite's 13-mer motif, sum the log2 weights
   over the atlas -5..+4 frame (the central S/T at index 6 is unscored) to get a
   per-kinase log-odds score.
3. ``percentile_by_kinase`` + ``build_kinome_net`` -- convert each kinase's
   scores to a within-cohort percentile and assign a site to a kinase when its
   percentile clears ``assign_percentile`` (Johnson's standard >=90th-percentile
   call). This substitutes the OSD-462 phosphoproteome itself for Johnson's
   Ochoa reference distribution, which is not redistributed on disk -- a
   documented, honest adaptation.
4. ``ksea`` -- the existing, independently tested KSEA statistic, unchanged.
5. ``over_representation`` -- a parent-gene-aware one-sided Fisher cross-check:
   are a kinase's predicted substrate *genes* enriched for suppressed sites?

Positive control: STLK3 (SPAK) and OSR1 must return negative KSEA z (inferred
activity down), consistent with NCC-activating-cluster suppression in OSD-462.
The wording stays prioritisation, not mechanism.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Mapping, Sequence

import numpy as np
import pandas as pd
from scipy import stats

from src.v11.core_analysis import bh
from src.multiomics.regulator_activity import (
    ksea,
    ksea_positive_control_passes,
    load_kinase_substrate_net,
)

# --------------------------------------------------------------------------- #
# On-disk locations
# --------------------------------------------------------------------------- #

ATLAS_PSSM_PATH = Path(
    "data/external/kinase_substrate/johnson2023_atlas/johnson2023_st_kinome_pssm.xlsx"
)
PSSM_SHEET = "ser_thr_all_norm_scaled_matrice"  # position-normalised, log2-ready

OSD462_PHOSPHO_XLSX = Path(
    "data/external/osdr/OSD-462/"
    "GLDS-462_phosphproteomics_2021-12-31_tc882-883_Pho_WorkUp_JM.xlsx"
)
PHOSPHO_SHEET = "siteQuant_360"
PHOSPHO_HEADER_ROW = 2          # 0-based; data begin on the next row
PHOSPHO_GENE_COL = 1            # gene_symbol
PHOSPHO_POS_COL = 3            # Site Position
PHOSPHO_MOTIF_COL = 4          # Motif (13-mer)
OSD462_ANCHOR_RUN = Path("data/results/run_20260522_osd462_anchor")

#: Effect/SE table written by the verified v9/OSD-462 anchor pipeline.
PHOSPHO_SITES_REL = "osd462_anchor/phospho_all_sites.tsv"

# --------------------------------------------------------------------------- #
# Atlas / motif geometry
# --------------------------------------------------------------------------- #

ATLAS_POSITIONS: tuple[int, ...] = (-5, -4, -3, -2, -1, 1, 2, 3, 4)
MOTIF_CENTER_INDEX = 6          # 13-mer: six residues either side of the site
CENTRAL_RESIDUES = ("S", "T")   # Ser/Thr atlas; Tyr sites are excluded
EPS = 1e-9

#: Atlas labels for the manuscript's positive-control axis. SPAK is recorded in
#: the Johnson atlas as ``STLK3``; OSR1 keeps its name; the WNK panel is present
#: as WNK1/WNK3/WNK4 (WNK2 is absent from the Ser/Thr atlas).
SPAK_OSR1_KINASES: tuple[str, ...] = ("STLK3", "OSR1")
WNK_KINASES: tuple[str, ...] = ("WNK1", "WNK3", "WNK4")

ASSIGN_PERCENTILE = 90.0        # Johnson's standard predicted-kinase floor
K_PER_SITE = 5                 # assign each site to its top-k best-matching kinases
MIN_SUBSTRATES = 3              # KSEA minimum quantified substrates per kinase

_HEADER_TOKEN = re.compile(r"^(-?\d+)([A-Za-z])$")


# --------------------------------------------------------------------------- #
# (1) atlas PSSMs
# --------------------------------------------------------------------------- #

class AtlasPSSM:
    """A (kinase x position x residue) tensor of log2 PSSM weights."""

    __slots__ = ("kinases", "positions", "residues", "logmat", "_res_ix", "_pos_ix")

    def __init__(
        self,
        kinases: Sequence[str],
        positions: Sequence[int],
        residues: Sequence[str],
        logmat: np.ndarray,
    ) -> None:
        self.kinases = list(kinases)
        self.positions = list(positions)
        self.residues = list(residues)
        self.logmat = logmat  # (K, P, R)
        self._res_ix = {r: i for i, r in enumerate(self.residues)}
        self._pos_ix = {p: i for i, p in enumerate(self.positions)}

    def __len__(self) -> int:
        return len(self.kinases)


def load_atlas_pssms(
    path: str | Path = ATLAS_PSSM_PATH,
    *,
    sheet: str = PSSM_SHEET,
    positions: Sequence[int] = ATLAS_POSITIONS,
) -> AtlasPSSM:
    """Read the Ser/Thr atlas into log2 weights.

    Column headers are ``{position}{residue}`` (e.g. ``-3R``); the
    position-normalised scaled matrix is positive everywhere, so ``log2`` of each
    cell is a clean per-position log-odds contribution that sums across the
    window.
    """
    import openpyxl  # lazy: only the loader needs it

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sh = wb[sheet]
        rows = sh.iter_rows(values_only=True)
        header = next(rows)
        # map each data column -> (position, residue)
        col_pos: list[int] = []
        col_res: list[str] = []
        col_keep: list[int] = []
        for j, tok in enumerate(header[1:], start=1):
            if tok is None:
                continue
            m = _HEADER_TOKEN.match(str(tok))
            if not m:
                continue
            p = int(m.group(1))
            if p not in positions:
                continue
            col_pos.append(p)
            col_res.append(m.group(2))
            col_keep.append(j)

        residues = sorted(set(col_res))
        pos_list = list(positions)
        pos_ix = {p: i for i, p in enumerate(pos_list)}
        res_ix = {r: i for i, r in enumerate(residues)}

        kinases: list[str] = []
        mats: list[np.ndarray] = []
        for row in rows:
            if row[0] is None:
                continue
            kinases.append(str(row[0]).strip())
            a = np.full((len(pos_list), len(residues)), np.nan, dtype=float)
            for j, p, r in zip(col_keep, col_pos, col_res):
                v = row[j]
                if v is None:
                    continue
                a[pos_ix[p], res_ix[r]] = float(v)
            mats.append(a)
    finally:
        wb.close()

    arr = np.stack(mats, axis=0)                       # (K, P, R)
    logmat = np.log2(np.clip(arr, EPS, None))
    return AtlasPSSM(kinases, pos_list, residues, logmat)


# --------------------------------------------------------------------------- #
# (2) motif scoring
# --------------------------------------------------------------------------- #

def _central_residue(motif: str, center: int = MOTIF_CENTER_INDEX) -> str | None:
    if motif is None:
        return None
    m = str(motif)
    return m[center] if 0 <= center < len(m) else None


def score_sites(
    motifs: Sequence[str],
    atlas: AtlasPSSM,
    *,
    center: int = MOTIF_CENTER_INDEX,
) -> np.ndarray:
    """Per-site, per-kinase summed log2 PSSM score over the atlas frame.

    ``motifs`` is a sequence of 13-mers centred on the phosphosite. Returns an
    ``(n_sites, n_kinases)`` array. Residues absent from a kinase's matrix (rare
    priming positions, terminal padding) contribute nothing for that position.
    """
    K = len(atlas)
    P = len(atlas.positions)
    N = len(motifs)
    logmat = atlas.logmat
    res_ix = atlas._res_ix

    # residue index per (position, site); -1 => skip
    ridx = np.full((P, N), -1, dtype=np.int64)
    for pi, p in enumerate(atlas.positions):
        mi = center + p
        for si in range(N):
            m = motifs[si]
            if m is None:
                continue
            if 0 <= mi < len(m):
                ridx[pi, si] = res_ix.get(m[mi], -1)

    scores = np.zeros((N, K), dtype=float)
    for pi in range(P):
        col = ridx[pi]
        valid = col >= 0
        if not valid.any():
            continue
        contrib = np.zeros((K, N), dtype=float)
        weights = logmat[:, pi, :]                     # (K, R)
        contrib[:, valid] = weights[:, col[valid]]
        # NaN weight (residue unseen for a kinase) -> no contribution
        np.nan_to_num(contrib, copy=False)
        scores += contrib.T
    return scores


def percentile_by_kinase(scores: np.ndarray) -> np.ndarray:
    """Within-cohort percentile of each site's score, computed per kinase.

    Column ``k`` is replaced by the percentile rank (0--100) of each site within
    the distribution of kinase ``k``'s scores across all sites -- the on-disk
    stand-in for Johnson's Ochoa reference distribution.
    """
    N, K = scores.shape
    if N <= 1:
        return np.full_like(scores, 100.0)
    pct = np.empty_like(scores)
    for k in range(K):
        r = stats.rankdata(scores[:, k], method="average")
        pct[:, k] = (r - 1.0) / (N - 1.0) * 100.0
    return pct


# --------------------------------------------------------------------------- #
# (3) kinome substrate net
# --------------------------------------------------------------------------- #

def assign_mask(
    percentile: np.ndarray,
    *,
    threshold: float = ASSIGN_PERCENTILE,
    k_per_site: int | None = K_PER_SITE,
) -> np.ndarray:
    """Boolean (n_sites x n_kinases) assignment mask.

    A site is assigned to a kinase when (a) its score clears the kinase's
    ``threshold`` percentile *and* (b) that kinase is among the site's
    ``k_per_site`` best-matching kinases. The per-site top-k constraint is what
    restores substrate specificity: because the percentile is computed *within
    cohort* (no external reference), a bare per-kinase floor would assign a fixed
    ``100 - threshold`` percent of sites to *every* kinase, collapsing the KSEA
    substrate sets to a near-constant size and an n-inflated z. Capping each site
    to its top-k kinases concentrates each kinase's net on its cognate motifs.
    """
    mask = percentile >= threshold
    N, K = percentile.shape
    if k_per_site and k_per_site < K:
        kth = K - k_per_site
        # indices of each site's top-k kinases by percentile
        top_idx = np.argpartition(percentile, kth=kth, axis=1)[:, kth:]
        topk = np.zeros_like(mask)
        np.put_along_axis(topk, top_idx, True, axis=1)
        mask &= topk
    return mask


def build_kinome_net(
    percentile: np.ndarray,
    atlas: AtlasPSSM,
    gene_symbol: Sequence[str],
    site_position: Sequence[object],
    *,
    threshold: float = ASSIGN_PERCENTILE,
    k_per_site: int | None = K_PER_SITE,
) -> pd.DataFrame:
    """Assign sites to their top-ranked kinases; emit the KSEA net schema.

    Emits exactly the ``kinase / substrate_gene / substrate_site`` schema that
    :func:`load_kinase_substrate_net` validates and :func:`ksea` consumes.
    """
    genes = list(gene_symbol)
    sites = [str(s) for s in site_position]
    mask = assign_mask(percentile, threshold=threshold, k_per_site=k_per_site)
    site_ix, kin_ix = np.nonzero(mask)
    kinases = atlas.kinases
    net = pd.DataFrame(
        {
            "kinase": [kinases[k] for k in kin_ix],
            "substrate_gene": [genes[i] for i in site_ix],
            "substrate_site": [sites[i] for i in site_ix],
        },
        columns=["kinase", "substrate_gene", "substrate_site"],
    )
    return net


def add_rank(ksea_table: pd.DataFrame) -> pd.DataFrame:
    """Add a most-suppressed-first rank over scored kinases (z ascending)."""
    out = ksea_table.reset_index(drop=True).copy()
    scored = out["status"].eq("scored")
    out["rank"] = np.nan
    out.loc[scored, "rank"] = np.arange(1, int(scored.sum()) + 1, dtype=float)
    return out


# --------------------------------------------------------------------------- #
# (4) parent-gene-aware over-representation cross-check
# --------------------------------------------------------------------------- #

def over_representation(
    sites: pd.DataFrame,
    net: pd.DataFrame,
    *,
    effect_col: str = "phospho_effect",
    p_col: str = "phospho_p_value",
    gene_col: str = "gene_symbol",
    p_thresh: float = 0.05,
    min_substrate_genes: int = MIN_SUBSTRATES,
) -> pd.DataFrame:
    """One-sided Fisher: are a kinase's substrate *genes* enriched for down sites?

    Sites are collapsed to one representative per parent gene (lowest phospho
    p-value, then most-negative effect) so a single gene with many sites cannot
    dominate. ``down`` := representative effect < 0 and p < ``p_thresh``.
    """
    g = sites.dropna(subset=[gene_col]).copy()
    g["_p"] = pd.to_numeric(g[p_col], errors="coerce")
    g["_e"] = pd.to_numeric(g[effect_col], errors="coerce")
    g = g.dropna(subset=["_e"])
    rep = (
        g.sort_values([gene_col, "_p", "_e"], ascending=[True, True, True])
        .groupby(gene_col, as_index=False)
        .head(1)
    )
    rep["is_down"] = (rep["_e"] < 0) & (rep["_p"] < p_thresh)
    all_genes = set(rep[gene_col])
    down_genes = set(rep.loc[rep["is_down"], gene_col])
    n_all = len(all_genes)
    n_down = len(down_genes)

    rows: list[dict[str, object]] = []
    for kin, grp in net.groupby("kinase"):
        ksub = set(grp["substrate_gene"]) & all_genes
        if len(ksub) < min_substrate_genes:
            continue
        a = len(ksub & down_genes)          # kinase-substrate & down
        b = len(ksub) - a                   # kinase-substrate & not-down
        c = n_down - a                      # not-substrate & down
        d = n_all - len(ksub) - c           # not-substrate & not-down
        odds, p = stats.fisher_exact([[a, b], [c, d]], alternative="greater")
        rows.append(
            {
                "kinase": kin,
                "n_substrate_genes": len(ksub),
                "n_down": a,
                "frac_down": a / len(ksub),
                "odds_ratio": float(odds),
                "p_value": float(p),
            }
        )
    out = pd.DataFrame(rows)
    if not out.empty:
        out["q_value"] = bh(out["p_value"].to_numpy(dtype=float))
        out = out.sort_values("p_value").reset_index(drop=True)
    return out


# --------------------------------------------------------------------------- #
# Inputs: join phospho effects to atlas motifs
# --------------------------------------------------------------------------- #

def _load_motif_map(
    xlsx: str | Path = OSD462_PHOSPHO_XLSX,
    *,
    sheet: str = PHOSPHO_SHEET,
) -> pd.DataFrame:
    import openpyxl  # lazy

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        sh = wb[sheet]
        rows = list(sh.iter_rows(values_only=True))
    finally:
        wb.close()
    recs = []
    for r in rows[PHOSPHO_HEADER_ROW + 1:]:
        gene, pos, motif = r[PHOSPHO_GENE_COL], r[PHOSPHO_POS_COL], r[PHOSPHO_MOTIF_COL]
        if gene is None or pos is None or motif is None:
            continue
        recs.append((str(gene).strip(), str(pos).strip(), str(motif).strip()))
    mp = pd.DataFrame(recs, columns=["gene_symbol", "site_position", "motif"])
    return mp.drop_duplicates(["gene_symbol", "site_position"])


def load_osd462_st_sites(
    root: str | Path,
    *,
    sites_root: str | Path | None = None,
    phospho_xlsx: str | Path = OSD462_PHOSPHO_XLSX,
    sheet: str = PHOSPHO_SHEET,
) -> pd.DataFrame:
    """Join the verified phospho effect table to atlas motifs; keep S/T sites.

    Returns ``gene_symbol, site_position, motif, phospho_effect, phospho_se,
    phospho_p_value`` for every quantified single phosphosite whose motif is
    centred on Ser or Thr (Tyr sites are dropped -- the atlas is Ser/Thr only).
    """
    root = Path(root)
    candidate_roots = [Path(sites_root)] if sites_root is not None else []
    candidate_roots.extend([root, OSD462_ANCHOR_RUN])
    sites_path = next((r / PHOSPHO_SITES_REL for r in candidate_roots if (r / PHOSPHO_SITES_REL).exists()), None)
    if sites_path is None:
        searched = [str(r / PHOSPHO_SITES_REL) for r in candidate_roots]
        raise FileNotFoundError(f"OSD-462 phosphosite effects not found; searched {searched}")
    eff = pd.read_csv(sites_path, sep="\t")
    eff["site_position"] = eff["site_position"].astype(str)
    eff["gene_symbol"] = eff["gene_symbol"].astype(str)

    motifs = _load_motif_map(phospho_xlsx, sheet=sheet)
    joined = eff.merge(motifs, on=["gene_symbol", "site_position"], how="inner")
    central = joined["motif"].map(_central_residue)
    keep = central.isin(CENTRAL_RESIDUES)
    out = joined.loc[keep].reset_index(drop=True)
    return out


# --------------------------------------------------------------------------- #
# Orchestrator
# --------------------------------------------------------------------------- #

def _panel_stats(ksea_table: pd.DataFrame, kinases: Sequence[str]) -> dict[str, object]:
    sub = ksea_table[ksea_table["kinase"].isin(list(kinases)) & ksea_table["status"].eq("scored")]
    if sub.empty:
        return {"present": [], "mean_z": float("nan"), "best_rank": None, "all_down": False}
    return {
        "present": sub["kinase"].tolist(),
        "mean_z": float(sub["z_score"].mean()),
        "best_rank": float(sub["rank"].min()),
        "all_down": bool((sub["z_score"] < 0).all()),
        "max_q": float(sub["q_value"].max()),
    }


def run_kinome_atlas_ksea(
    root: str | Path,
    *,
    sites_root: str | Path | None = None,
    assign_percentile: float = ASSIGN_PERCENTILE,
    k_per_site: int = K_PER_SITE,
    min_substrates: int = MIN_SUBSTRATES,
    atlas_path: str | Path = ATLAS_PSSM_PATH,
    phospho_xlsx: str | Path = OSD462_PHOSPHO_XLSX,
) -> dict[str, object]:
    """End-to-end Repair A: atlas -> motif scores -> kinome net -> KSEA + over-rep."""
    root = Path(root)
    out_dir = root / "regulator_activity"
    out_dir.mkdir(parents=True, exist_ok=True)

    sites = load_osd462_st_sites(root, sites_root=sites_root, phospho_xlsx=phospho_xlsx)
    atlas = load_atlas_pssms(atlas_path)

    scores = score_sites(sites["motif"].tolist(), atlas)
    pct = percentile_by_kinase(scores)
    net = build_kinome_net(
        pct, atlas, sites["gene_symbol"].tolist(), sites["site_position"].tolist(),
        threshold=assign_percentile, k_per_site=k_per_site,
    )

    ksea_sites = sites[["gene_symbol", "site_position", "phospho_effect"]].copy()
    ksea_table = ksea(ksea_sites, net, min_substrates=min_substrates)
    ksea_table = add_rank(ksea_table)

    overrep = over_representation(sites, net, min_substrate_genes=min_substrates)

    # ---- artifacts -------------------------------------------------------- #
    ksea_path = out_dir / "osd462_kinome_atlas_ksea.tsv"
    overrep_path = out_dir / "osd462_kinome_atlas_overrep.tsv"
    net_summary_path = out_dir / "osd462_kinome_atlas_net_summary.tsv"
    ksea_table.to_csv(ksea_path, sep="\t", index=False)
    overrep.to_csv(overrep_path, sep="\t", index=False)
    net.groupby("kinase").size().rename("n_substrate_sites").reset_index().to_csv(
        net_summary_path, sep="\t", index=False
    )

    spak_osr1 = _panel_stats(ksea_table, SPAK_OSR1_KINASES)
    wnk = _panel_stats(ksea_table, WNK_KINASES)
    control_pass = ksea_positive_control_passes(ksea_table, control_kinases=SPAK_OSR1_KINASES)

    n_scored = int(ksea_table["status"].eq("scored").sum())
    verdict = {
        "analysis": "repair_a_kinome_atlas_ksea",
        "n_sites_scored": int(len(sites)),
        "n_kinases_total": int(len(atlas)),
        "n_kinases_scored": n_scored,
        "assign_percentile": float(assign_percentile),
        "k_per_site": int(k_per_site),
        "median_substrates_per_kinase": float(
            ksea_table.loc[ksea_table["status"].eq("scored"), "n_substrates_quantified"].median()
        ) if n_scored else float("nan"),
        "min_substrates": int(min_substrates),
        "spak_osr1_panel": spak_osr1,
        "wnk_panel": wnk,
        "positive_control_passes": bool(control_pass),
        # ---- headline-index provenance keys ----
        "kinome_spak_osr1_z": spak_osr1["mean_z"],
        "kinome_spak_osr1_rank": spak_osr1["best_rank"],
        "kinome_wnk_z": wnk["mean_z"],
        "outputs": {
            "ksea": str(ksea_path),
            "over_representation": str(overrep_path),
            "net_summary": str(net_summary_path),
        },
    }
    (out_dir / "osd462_kinome_atlas_verdict.json").write_text(json.dumps(verdict, indent=2))
    return verdict


def main() -> None:
    ap = argparse.ArgumentParser(description="Repair A -- kinome-wide atlas KSEA (OSD-462)")
    ap.add_argument("--run-root", required=True, help="results run directory (contains osd462_anchor/)")
    ap.add_argument("--sites-root", default=None,
                    help="optional run directory containing osd462_anchor/phospho_all_sites.tsv")
    ap.add_argument("--assign-percentile", type=float, default=ASSIGN_PERCENTILE,
                    help="per-kinase percentile floor for assignment (default 90)")
    ap.add_argument("--k-per-site", type=int, default=K_PER_SITE,
                    help="assign each site to its top-k best-matching kinases (default 5)")
    ap.add_argument("--min-substrates", type=int, default=MIN_SUBSTRATES)
    args = ap.parse_args()
    verdict = run_kinome_atlas_ksea(
        args.run_root,
        sites_root=args.sites_root,
        assign_percentile=args.assign_percentile,
        k_per_site=args.k_per_site,
        min_substrates=args.min_substrates,
    )
    print(json.dumps(verdict, indent=2))


if __name__ == "__main__":
    main()
