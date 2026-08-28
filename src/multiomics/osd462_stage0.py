"""Reproducible Stage 0 provenance checks for the OSD-462 MS assays.

This module deliberately separates three objects that were previously mixed in
the analysis narrative:

* biological samples and their reporter channels;
* multiplex-level raw LC-MS acquisitions (which are not sample-specific);
* workbook phosphosite features and their residue identities.

The functions read the source workbooks and ISA metadata directly.  They do
not use manuscript tables or previously generated analysis outputs.
"""
from __future__ import annotations

import hashlib
import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd


REPO_ROOT = Path(__file__).resolve().parents[2]
OSD462_ROOT = REPO_ROOT / "data" / "external" / "osdr" / "OSD-462"
METADATA_ROOT = OSD462_ROOT / "metadata"

PROTEIN_WORKBOOK = (
    OSD462_ROOT
    / "GLDS-462_proteomics_2021-12-31_tc884-885_Protein_WorkUp.xlsx"
)
PHOSPHO_WORKBOOK = (
    OSD462_ROOT
    / "GLDS-462_phosphproteomics_2021-12-31_tc882-883_Pho_WorkUp_JM.xlsx"
)
STUDY_FILE = METADATA_ROOT / "s_OSD-462.txt"
INVESTIGATION_FILE = METADATA_ROOT / "i_Investigation.txt"
PROTEIN_ASSAY_FILE = (
    METADATA_ROOT
    / "a_OSD-462_protein-expression-profiling_mass-spectrometry_orbitrap-fusion-lumos.txt"
)
PHOSPHO_ASSAY_FILE = (
    METADATA_ROOT
    / "a_OSD-462_phosphoprotein-profiling_mass-spectrometry_orbitrap-eclipse.txt"
)

PLEXES = ("Samp1-5", "Samp6-10")
CONDITION_LABELS = {
    "BL": "Basal Control",
    "FL": "Space Flight",
    "GC": "Ground Control",
}

# Residue identity is part of the key.  This prevents a position-only feature
# such as NCC p65 from being silently treated as a serine/threonine substrate.
STRICT_RENAL_SITES: dict[tuple[str, int, str], dict[str, str]] = {
    ("Slc12a3", 53, "T"): {
        "axis": "SPAK_OSR1_to_NCC",
        "role": "canonical_NCC_activating_site",
        "evidence_id": "PMID:24039833",
    },
    ("Slc12a3", 58, "T"): {
        "axis": "SPAK_OSR1_to_NCC",
        "role": "canonical_NCC_activating_site",
        "evidence_id": "PMID:24039833",
    },
    ("Slc12a3", 71, "S"): {
        "axis": "SPAK_OSR1_to_NCC",
        "role": "canonical_NCC_activating_site",
        "evidence_id": "PMID:24039833",
    },
    ("Stk39", 243, "T"): {
        "axis": "WNK_to_SPAK",
        "role": "canonical_SPAK_T_loop_activation_site",
        "evidence_id": "PMCID:PMC2944316",
    },
    ("Stk39", 383, "S"): {
        "axis": "WNK_to_SPAK",
        "role": "canonical_SPAK_S_motif_regulatory_site",
        "evidence_id": "PMCID:PMC2944316",
    },
}


def sha256_file(path: str | Path) -> str:
    """Return the SHA-256 digest of a file."""
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _clean_plex(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip().strip('"')
    return text if text in PLEXES else None


def _normalise_tag(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    return str(value).strip().upper()


def _normalise_position(value: object) -> str:
    """Normalise workbook site positions while preserving composites."""
    if value is None:
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    pieces = []
    for piece in str(value).strip().split(";"):
        piece = piece.strip()
        try:
            numeric = float(piece)
            pieces.append(str(int(numeric)) if numeric.is_integer() else piece)
        except ValueError:
            pieces.append(piece)
    return ";".join(pieces)


def _workbook_to_isa_sample(workbook_sample: str) -> str:
    match = re.fullmatch(r"RR-10_(BL|FL|GC)-0*(\d+)", workbook_sample)
    if not match:
        raise ValueError(f"unrecognised OSD-462 workbook sample: {workbook_sample!r}")
    condition, animal = match.groups()
    isa_condition = {"BL": "BSL", "FL": "FLT", "GC": "GC"}[condition]
    animal_prefix = {"BL": "B", "FL": "F", "GC": "G"}[condition]
    return f"RR10_KDN_WT_{isa_condition}_{animal_prefix}{int(animal)}"


def extract_intro_design(
    workbook: str | Path,
    *,
    modality: str,
) -> pd.DataFrame:
    """Read the exact sample-to-plex-to-reporter map from ``INTRO-DATA``."""
    import openpyxl

    workbook = Path(workbook)
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["INTRO-DATA"]
    records: list[dict[str, object]] = []
    current_plex: str | None = None
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row = tuple(row)
        maybe_plex = _clean_plex(row[0] if len(row) > 0 else None)
        if maybe_plex:
            current_plex = maybe_plex
        sample = row[1] if len(row) > 1 else None
        tag = row[2] if len(row) > 2 else None
        if not isinstance(sample, str) or not sample.startswith("RR-10_"):
            continue
        if current_plex is None:
            raise ValueError(f"sample {sample!r} appears before a plex label")
        match = re.fullmatch(r"RR-10_(BL|FL|GC)-0*(\d+)", sample.strip())
        if not match:
            raise ValueError(f"unrecognised sample label in {workbook.name}: {sample!r}")
        condition, animal = match.groups()
        records.append(
            {
                "modality": modality,
                "plex": current_plex,
                "workbook_sample_name": sample.strip(),
                "condition_code": condition,
                "condition": CONDITION_LABELS[condition],
                "animal_number": int(animal),
                "reporter_tag": _normalise_tag(tag),
                "workbook_intro_excel_row": excel_row,
                "source_workbook": str(workbook.relative_to(REPO_ROOT)),
            }
        )
    wb.close()
    out = pd.DataFrame(records)
    if len(out) != 30:
        raise ValueError(f"expected 30 samples in {workbook.name}; found {len(out)}")
    return out


def extract_raw_inventory(
    workbook: str | Path,
    *,
    modality: str,
) -> pd.DataFrame:
    """Read multiplex-level raw acquisition IDs and aliases from a workbook."""
    import openpyxl

    workbook = Path(workbook)
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    sheet = "RAWfiles" if modality == "protein" else "RAW files"
    ws = wb[sheet]
    records: list[dict[str, object]] = []

    if modality == "protein":
        groups = (
            ("Samp1-5", 1, 2, None, 3),
            ("Samp6-10", 5, 6, None, 3),
        )
    elif modality == "phosphoprotein":
        groups = (
            ("Samp1-5", 1, 2, 3, 4),
            ("Samp6-10", 6, 7, 8, 4),
        )
    else:
        raise ValueError(f"unsupported modality: {modality!r}")

    all_rows = list(ws.iter_rows(values_only=True))
    for plex, raw_idx, alias_idx, cv_idx, first_excel_row in groups:
        for excel_row, row in enumerate(all_rows, start=1):
            if excel_row < first_excel_row:
                continue
            raw_id = row[raw_idx] if len(row) > raw_idx else None
            alias = row[alias_idx] if len(row) > alias_idx else None
            if raw_id is None or str(raw_id).strip() == "":
                continue
            cv = row[cv_idx] if cv_idx is not None and len(row) > cv_idx else None
            records.append(
                {
                    "modality": modality,
                    "plex": plex,
                    "raw_acquisition_id": str(raw_id).strip(),
                    "raw_alias": "" if alias is None else str(alias).strip(),
                    "compensation_voltage": "" if cv is None else str(cv).strip(),
                    "workbook_raw_excel_row": excel_row,
                    "source_workbook": str(workbook.relative_to(REPO_ROOT)),
                    "raw_files_shared_across_plex": True,
                    "raw_file_sample_specific": False,
                }
            )
    wb.close()
    return pd.DataFrame(records)


def _read_isa_metadata() -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    study = pd.read_csv(STUDY_FILE, sep="\t", dtype=str, keep_default_na=False)
    assays = {
        "protein": pd.read_csv(
            PROTEIN_ASSAY_FILE, sep="\t", dtype=str, keep_default_na=False
        ),
        "phosphoprotein": pd.read_csv(
            PHOSPHO_ASSAY_FILE, sep="\t", dtype=str, keep_default_na=False
        ),
    }
    return study, assays


def build_sample_design() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build the exact sample design and multiplex-level raw inventory.

    Returns
    -------
    design
        One row per biological sample per MS modality (60 rows).
    raw_inventory
        One row per multiplex-level raw LC-MS acquisition (96 rows).
    """
    modality_workbooks = {
        "protein": PROTEIN_WORKBOOK,
        "phosphoprotein": PHOSPHO_WORKBOOK,
    }
    study, assays = _read_isa_metadata()
    design_parts: list[pd.DataFrame] = []
    raw_parts: list[pd.DataFrame] = []

    # Exact biological-sample rows only; technical-replicate RNA rows in the
    # study file are intentionally excluded.
    study_exact = study[~study["Sample Name"].str.contains("_techrep", regex=False)].copy()
    study_exact = study_exact.drop_duplicates("Sample Name")

    for modality, workbook in modality_workbooks.items():
        design = extract_intro_design(workbook, modality=modality)
        design["isa_sample_name"] = design["workbook_sample_name"].map(
            _workbook_to_isa_sample
        )
        assay = assays[modality].copy()
        assay = assay.rename(
            columns={
                "Sample Name": "isa_sample_name",
                "Label": "isa_reporter_tag",
                "Raw Spectral Data File": "raw_archive",
                "Protein Assignment File": "isa_processed_workbook",
                "Parameter Value[Instrument]": "instrument",
                "Parameter Value[Dissociation]": "dissociation",
            }
        )
        assay_keep = [
            "isa_sample_name",
            "isa_reporter_tag",
            "raw_archive",
            "isa_processed_workbook",
            "instrument",
            "dissociation",
        ]
        design = design.merge(assay[assay_keep], on="isa_sample_name", how="left")

        study_meta = study_exact.rename(
            columns={
                "Sample Name": "isa_sample_name",
                "Characteristics[Strain]": "strain",
                "Characteristics[Genotype]": "genotype",
                "Characteristics[Sex]": "sex",
                "Characteristics[Material Type]": "tissue",
                "Characteristics[Age at Launch]": "age_at_launch",
                "Factor Value[Spaceflight]": "isa_condition",
                "Source Name": "isa_source_name",
            }
        )
        meta_keep = [
            "isa_sample_name",
            "isa_source_name",
            "strain",
            "genotype",
            "sex",
            "tissue",
            "age_at_launch",
            "isa_condition",
        ]
        design = design.merge(study_meta[meta_keep], on="isa_sample_name", how="left")
        design["isa_reporter_tag"] = design["isa_reporter_tag"].map(_normalise_tag)
        design["reporter_tag_matches_isa"] = (
            design["reporter_tag"] == design["isa_reporter_tag"]
        )
        design["workbook_sample_matches_isa_source"] = (
            design["workbook_sample_name"] == design["isa_source_name"]
        )
        design["analysis_primary_flight_vs_ground"] = design["condition_code"].isin(
            ["FL", "GC"]
        )
        # Preserve protocol evidence and contradictory legacy metadata as
        # separate fields; do not collapse them into one ambiguous assay label.
        design["resolved_assay_name"] = "TMTpro isobaric-tag proteomics"
        design["detailed_protocol_labeling_reagent"] = "TMTpro"
        design["detailed_protocol_chemistry_evidence"] = (
            "TMTpro reagent; +304.207-Da tag mass; TMTpro impurity correction"
        )
        design["legacy_investigation_assay_term"] = "iTRAQ"
        design["assay_metadata_discrepancy"] = True

        raw = extract_raw_inventory(workbook, modality=modality)
        archive_by_plex = (
            design.groupby("plex")["raw_archive"]
            .agg(lambda x: ";".join(sorted(set(x))))
            .to_dict()
        )
        raw["raw_archive"] = raw["plex"].map(archive_by_plex)
        raw_parts.append(raw)

        raw_summary = (
            raw.groupby("plex")
            .agg(
                n_raw_acquisitions_in_plex=("raw_acquisition_id", "size"),
                raw_acquisition_ids=(
                    "raw_acquisition_id",
                    lambda x: ";".join(x.astype(str)),
                ),
                raw_aliases=("raw_alias", lambda x: ";".join(x.astype(str))),
            )
            .reset_index()
        )
        design = design.merge(raw_summary, on="plex", how="left")
        design["raw_files_shared_across_plex"] = True
        design["raw_file_sample_specific"] = False
        design_parts.append(design)

    out = pd.concat(design_parts, ignore_index=True)
    raw_out = pd.concat(raw_parts, ignore_index=True)

    # A condition has a unique reporter set, reused in the second plex with no
    # label swap.  Store this design property explicitly on every row.
    tag_condition = out[["reporter_tag", "condition_code"]].drop_duplicates()
    perfect_tag_mapping = bool(
        tag_condition.groupby("reporter_tag")["condition_code"].nunique().max() == 1
    )
    mapping_by_plex = {
        plex: set(
            map(
                tuple,
                out[out["plex"].eq(plex)][["reporter_tag", "condition_code"]]
                .drop_duplicates()
                .to_numpy(),
            )
        )
        for plex in PLEXES
    }
    no_label_swap = mapping_by_plex[PLEXES[0]] == mapping_by_plex[PLEXES[1]]
    out["condition_perfectly_confounded_with_reporter_tag"] = (
        perfect_tag_mapping and no_label_swap
    )
    out["reporter_assignment_swapped_between_plexes"] = not no_label_swap

    order = [
        "modality",
        "plex",
        "workbook_sample_name",
        "isa_sample_name",
        "condition_code",
        "condition",
        "animal_number",
        "reporter_tag",
        "isa_reporter_tag",
        "reporter_tag_matches_isa",
        "workbook_sample_matches_isa_source",
        "genotype",
        "strain",
        "sex",
        "tissue",
        "age_at_launch",
        "analysis_primary_flight_vs_ground",
        "raw_archive",
        "raw_files_shared_across_plex",
        "raw_file_sample_specific",
        "n_raw_acquisitions_in_plex",
        "raw_acquisition_ids",
        "raw_aliases",
        "instrument",
        "dissociation",
        "resolved_assay_name",
        "detailed_protocol_labeling_reagent",
        "detailed_protocol_chemistry_evidence",
        "legacy_investigation_assay_term",
        "assay_metadata_discrepancy",
        "condition_perfectly_confounded_with_reporter_tag",
        "reporter_assignment_swapped_between_plexes",
        "source_workbook",
        "workbook_intro_excel_row",
        "isa_processed_workbook",
    ]
    return out[order].sort_values(
        ["modality", "plex", "condition_code", "animal_number"]
    ).reset_index(drop=True), raw_out.sort_values(
        ["modality", "plex", "workbook_raw_excel_row"]
    ).reset_index(drop=True)


def _split_components(value: object) -> list[str]:
    if value is None:
        return []
    return [piece.strip() for piece in str(value).split(";") if piece.strip()]


def motif_centre_residue(motif: str) -> str:
    """Return the central residue of a 13-aa localization motif."""
    motif = str(motif).strip().upper()
    if len(motif) != 13:
        return ""
    return motif[6]


def derive_site_components(
    site_position: object,
    motif_value: object,
) -> list[tuple[int, str, str]]:
    """Return ``(position, residue, motif)`` components for a workbook row."""
    positions = _split_components(_normalise_position(site_position))
    motifs = _split_components(motif_value)
    if len(positions) != len(motifs):
        raise ValueError(
            f"position/motif component mismatch: {positions!r} vs {motifs!r}"
        )
    components: list[tuple[int, str, str]] = []
    for position, motif in zip(positions, motifs):
        try:
            integer_position = int(position)
        except ValueError as exc:
            raise ValueError(f"non-integer site position: {position!r}") from exc
        residue = motif_centre_residue(motif)
        if residue not in {"S", "T", "Y"}:
            raise ValueError(
                f"site motif does not have S/T/Y at its centre: {motif!r}"
            )
        components.append((integer_position, residue, motif))
    return components


def map_sequence_phosphoform(
    site_position: object,
    motif_value: object,
    sequence_value: object,
) -> dict[str, object]:
    """Map every ``#`` in a workbook peptide sequence to an absolute site.

    The workbook encodes phosphorylation by placing ``#`` after the modified
    residue.  A 13-aa localization motif anchors its center to the reported
    position.  Longest-overlap alignment between each motif and the peptide
    determines the peptide's absolute start; independent component anchors
    must agree.
    """
    components = derive_site_components(site_position, motif_value)
    sequence = str(sequence_value).strip().upper()
    split = sequence.split(".")
    core = split[1] if len(split) >= 3 else sequence
    unmodified = core.replace("#", "")

    starts: list[int] = []
    overlaps: list[int] = []
    for position, _residue, motif in components:
        match = SequenceMatcher(None, motif, unmodified, autojunk=False).find_longest_match(
            0, len(motif), 0, len(unmodified)
        )
        if match.size < 5:
            return {
                "peptide_sequence_core": core,
                "peptide_sequence_unmodified": unmodified,
                "sequence_phosphoform_site_labels": "",
                "sequence_phosphoform_n_sites": core.count("#"),
                "sequence_phosphoform_mapping_status": "insufficient_motif_peptide_overlap",
                "sequence_anchor_overlap_lengths": ";".join(map(str, overlaps)),
            }
        motif_absolute_start = position - 6
        peptide_absolute_start = motif_absolute_start + match.a - match.b
        starts.append(peptide_absolute_start)
        overlaps.append(match.size)
    if len(set(starts)) != 1:
        return {
            "peptide_sequence_core": core,
            "peptide_sequence_unmodified": unmodified,
            "sequence_phosphoform_site_labels": "",
            "sequence_phosphoform_n_sites": core.count("#"),
            "sequence_phosphoform_mapping_status": "inconsistent_component_anchors",
            "sequence_anchor_overlap_lengths": ";".join(map(str, overlaps)),
        }

    peptide_start = starts[0]
    labels: list[str] = []
    peptide_index = 0
    previous_residue = ""
    previous_index = -1
    for character in core:
        if character == "#":
            if previous_residue and previous_index >= 0:
                labels.append(f"{previous_residue}{peptide_start + previous_index}")
            continue
        if character.isalpha():
            previous_residue = character
            previous_index = peptide_index
            peptide_index += 1
    status = "mapped" if len(labels) == core.count("#") else "unmapped_hash_marker"
    return {
        "peptide_sequence_core": core,
        "peptide_sequence_unmodified": unmodified,
        "sequence_phosphoform_site_labels": ";".join(labels),
        "sequence_phosphoform_n_sites": len(labels),
        "sequence_phosphoform_mapping_status": status,
        "sequence_anchor_overlap_lengths": ";".join(map(str, overlaps)),
    }


def _as_float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return np.nan


def _scaled_channel_columns(
    headers: Iterable[object],
    sample_labels: Iterable[object],
) -> list[dict[str, object]]:
    channels: list[dict[str, object]] = []
    for index, (header, sample) in enumerate(zip(headers, sample_labels)):
        if not isinstance(header, str):
            continue
        lower = header.lower()
        if "~" not in header or "scaled" not in lower or "_sn" not in lower:
            continue
        plex = header.split("~", 1)[0].strip()
        if plex not in PLEXES or not isinstance(sample, str):
            continue
        match = re.fullmatch(r"(BL|FL|GC)-0*(\d+)", sample.strip())
        if not match:
            continue
        channels.append(
            {
                "index": index,
                "plex": plex,
                "condition_code": match.group(1),
                "sample": sample.strip(),
            }
        )
    return channels


def audit_ncc_spak_phosphosites(
    workbook: str | Path = PHOSPHO_WORKBOOK,
) -> pd.DataFrame:
    """Audit every NCC and SPAK phosphosite feature from the source workbook."""
    import openpyxl

    workbook = Path(workbook)
    targets = {"Slc12a3", "Stk39"}
    configs = (
        {
            "sheet": "siteQuant_360",
            "kind": "single_site_rollup",
            "gene_col": "gene_symbol",
            "protein_col": "Protein Id",
            "position_col": "Site Position",
            "motif_col": "Motif",
            "score_col": "Max Score",
            "redundancy_col": "redundancy",
            "sequence_col": "sequence",
        },
        {
            "sheet": "siteQuant_360_compositeSite",
            "kind": "composite_site_rollup",
            "gene_col": "geneSymbol",
            "protein_col": "proteinID",
            "position_col": "sitePosStr",
            "motif_col": "motifPeptideStr",
            "score_col": "maxScoreStr",
            "redundancy_col": "redundancyStr",
            "sequence_col": "sequence",
        },
    )
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    records: list[dict[str, object]] = []
    for config in configs:
        ws = wb[config["sheet"]]
        row_iter = ws.iter_rows(values_only=True)
        row1 = next(row_iter)
        row2 = next(row_iter)
        headers = next(row_iter)
        del row1
        header_index = {
            str(value).strip(): index
            for index, value in enumerate(headers)
            if isinstance(value, str)
        }
        channels = _scaled_channel_columns(headers, row2)
        required = [
            config["gene_col"],
            config["protein_col"],
            config["position_col"],
            config["motif_col"],
            config["score_col"],
            config["redundancy_col"],
            config["sequence_col"],
            "Samp1-5~num_quant",
            "Samp6-10~num_quant",
        ]
        missing = [column for column in required if column not in header_index]
        if missing:
            raise KeyError(f"{config['sheet']} missing columns: {missing}")

        for excel_row, row in enumerate(row_iter, start=4):
            gene = row[header_index[config["gene_col"]]]
            if gene not in targets:
                continue
            position_string = _normalise_position(
                row[header_index[config["position_col"]]]
            )
            motif_string = row[header_index[config["motif_col"]]]
            components = derive_site_components(position_string, motif_string)
            score_parts = _split_components(row[header_index[config["score_col"]]])
            scores = [_as_float(value) for value in score_parts]
            if len(scores) not in {0, len(components)}:
                raise ValueError(
                    f"score/component mismatch at {config['sheet']} row {excel_row}"
                )

            strict = [
                STRICT_RENAL_SITES.get((str(gene), position, residue))
                for position, residue, _motif in components
            ]
            canonical_labels = [
                f"{residue}{position}"
                for (position, residue, _motif), annotation in zip(
                    components, strict
                )
                if annotation is not None
            ]
            noncanonical_labels = [
                f"{residue}{position}"
                for (position, residue, _motif), annotation in zip(
                    components, strict
                )
                if annotation is None
            ]
            all_canonical = bool(strict) and all(item is not None for item in strict)
            any_canonical = any(item is not None for item in strict)
            positive_counts = {
                "n_positive_scaled_all": 0,
                "n_positive_scaled_BL": 0,
                "n_positive_scaled_FL": 0,
                "n_positive_scaled_GC": 0,
                "n_positive_scaled_Samp1-5": 0,
                "n_positive_scaled_Samp6-10": 0,
            }
            for channel in channels:
                value = _as_float(row[int(channel["index"])])
                if np.isfinite(value) and value > 0:
                    positive_counts["n_positive_scaled_all"] += 1
                    positive_counts[
                        f"n_positive_scaled_{channel['condition_code']}"
                    ] += 1
                    positive_counts[f"n_positive_scaled_{channel['plex']}"] += 1

            protein_id = str(row[header_index[config["protein_col"]]]).strip()
            accession_match = re.match(r"(?:sp|tr)\|([^|]+)\|", protein_id)
            sequence = str(row[header_index[config["sequence_col"]]]).strip()
            sequence_mapping = map_sequence_phosphoform(
                position_string, motif_string, sequence
            )
            n_hash_modifications = sequence.count("#")
            feature_labels = {
                f"{residue}{position}"
                for position, residue, _motif in components
            }
            phosphoform_labels = set(
                _split_components(
                    sequence_mapping["sequence_phosphoform_site_labels"]
                )
            )
            isolated_canonical_feature = (
                all_canonical
                and config["kind"] == "single_site_rollup"
                and len(components) == 1
                and sequence_mapping["sequence_phosphoform_mapping_status"] == "mapped"
                and phosphoform_labels == feature_labels
            )
            if isolated_canonical_feature:
                claim_class = "isolated_strict_canonical_assay_feature"
            elif all_canonical and phosphoform_labels.issuperset(feature_labels):
                claim_class = "canonical_site_indexed_but_comodified_context_only"
            elif all_canonical:
                claim_class = "canonical_component_but_phosphoform_unresolved"
            elif any_canonical:
                claim_class = "mixed_composite_not_attributable_to_canonical_site"
            else:
                claim_class = "not_in_strict_canonical_set"
            record = {
                "gene_symbol": str(gene),
                "protein_id": protein_id,
                "uniprot_accession": (
                    accession_match.group(1) if accession_match else ""
                ),
                "source_sheet": config["sheet"],
                "site_feature_kind": config["kind"],
                "source_excel_row": excel_row,
                "workbook_site_position": position_string,
                "residue_resolved_site_labels": ";".join(
                    f"{residue}{position}"
                    for position, residue, _motif in components
                ),
                "component_residues": ";".join(
                    residue for _position, residue, _motif in components
                ),
                "motifs": ";".join(motif for _position, _residue, motif in components),
                "workbook_max_score_components": ";".join(score_parts),
                "workbook_redundancy_components": str(
                    row[header_index[config["redundancy_col"]]]
                ),
                "workbook_score_type_from_protocol": "AScore",
                "localization_score_protocol_threshold": 13,
                "localization_score_protocol_interpretation": (
                    "AScore >=13 denotes 95% confidence for a specified site"
                ),
                "all_localization_scores_ge_13": (
                    bool(scores)
                    and all(np.isfinite(score) and score >= 13 for score in scores)
                ),
                "sequence": sequence,
                "n_hash_modifications_in_sequence": n_hash_modifications,
                **sequence_mapping,
                "sequence_has_additional_modifications_beyond_feature_components": (
                    phosphoform_labels.issuperset(feature_labels)
                    and phosphoform_labels != feature_labels
                ),
                "feature_components_present_in_sequence_phosphoform": (
                    feature_labels.issubset(phosphoform_labels)
                ),
                "isolated_canonical_assay_feature": isolated_canonical_feature,
                "residue_indexed_canonical_but_comodified": (
                    all_canonical and not isolated_canonical_feature
                ),
                "n_quantified_peptides_Samp1-5": _as_float(
                    row[header_index["Samp1-5~num_quant"]]
                ),
                "n_quantified_peptides_Samp6-10": _as_float(
                    row[header_index["Samp6-10~num_quant"]]
                ),
                "explicit_row_level_psm_count_available": False,
                "quantification_protocol_note": (
                    "reporter signal summed across matching PSMs; explicit "
                    "row-level PSM count not supplied in this site-rollup sheet"
                ),
                "strict_canonical_components": ";".join(canonical_labels),
                "noncanonical_components": ";".join(noncanonical_labels),
                "any_component_strict_canonical": any_canonical,
                "all_components_strict_canonical": all_canonical,
                "canonical_claim_class": claim_class,
                "contains_Y65_or_Y68": any(
                    position in {65, 68} and residue == "Y"
                    for position, residue, _motif in components
                ),
                "Y65_Y68_canonical_SPAK_OSR1_site": False,
                "source_workbook": str(workbook.relative_to(REPO_ROOT)),
                **positive_counts,
            }
            records.append(record)
    wb.close()
    return pd.DataFrame(records).sort_values(
        ["gene_symbol", "site_feature_kind", "workbook_site_position"]
    ).reset_index(drop=True)


def build_stage0_qc(
    design: pd.DataFrame,
    raw_inventory: pd.DataFrame,
    phosphosite_audit: pd.DataFrame,
) -> pd.DataFrame:
    """Create machine-readable PASS/WARN/FAIL Stage 0 checks."""
    rows: list[dict[str, str]] = []

    def add(metric: str, status: str, value: object, detail: str) -> None:
        rows.append(
            {
                "metric": metric,
                "status": status,
                "value": str(value),
                "detail": detail,
            }
        )

    sample_counts = design.groupby("modality").size().to_dict()
    add(
        "sample_count_per_modality",
        "PASS" if set(sample_counts.values()) == {30} else "FAIL",
        sample_counts,
        "Expected 30 WT female kidney samples in each MS modality.",
    )
    primary_counts = (
        design[design["analysis_primary_flight_vs_ground"]]
        .groupby("modality")
        .size()
        .to_dict()
    )
    add(
        "flight_ground_count_per_modality",
        "PASS" if set(primary_counts.values()) == {20} else "FAIL",
        primary_counts,
        "Primary FL-vs-GC contrast has 10 flight and 10 ground samples.",
    )
    condition_counts = (
        design.groupby(["modality", "plex", "condition_code"])
        .size()
        .to_dict()
    )
    add(
        "condition_balance_within_plex",
        "PASS" if set(condition_counts.values()) == {5} else "FAIL",
        condition_counts,
        "Each plex contains five BL, five FL, and five GC samples.",
    )
    add(
        "workbook_isa_sample_and_tag_match",
        (
            "PASS"
            if design[
                ["reporter_tag_matches_isa", "workbook_sample_matches_isa_source"]
            ]
            .all()
            .all()
            else "FAIL"
        ),
        int(
            design[
                ["reporter_tag_matches_isa", "workbook_sample_matches_isa_source"]
            ]
            .all(axis=1)
            .sum()
        ),
        "All 60 sample-modality rows should agree between workbook and ISA metadata.",
    )
    biological_values = {
        column: sorted(design[column].dropna().astype(str).unique())
        for column in ["genotype", "strain", "sex", "tissue"]
    }
    expected_biology = (
        biological_values["genotype"] == ["Wild Type"]
        and biological_values["strain"] == ["B6129SF2/J"]
        and biological_values["sex"] == ["Female"]
        and biological_values["tissue"] == ["Left kidney"]
    )
    add(
        "biological_sample_scope",
        "PASS" if expected_biology else "FAIL",
        biological_values,
        "The MS comparison is WT female B6129SF2/J left kidney only.",
    )
    modality_maps = {
        modality: set(
            map(
                tuple,
                group[
                    ["workbook_sample_name", "plex", "reporter_tag"]
                ].to_numpy(),
            )
        )
        for modality, group in design.groupby("modality")
    }
    add(
        "protein_phosphoprotein_channel_map_identity",
        (
            "PASS"
            if modality_maps.get("protein") == modality_maps.get("phosphoprotein")
            else "FAIL"
        ),
        modality_maps.get("protein") == modality_maps.get("phosphoprotein"),
        "Protein and phosphoprotein workbooks assign the same samples to the same tags.",
    )
    raw_counts = raw_inventory.groupby(["modality", "plex"]).size().to_dict()
    add(
        "raw_acquisition_count_per_plex",
        "PASS" if set(raw_counts.values()) == {24} else "FAIL",
        raw_counts,
        "Each modality/plex has 24 multiplex-level LC-MS acquisitions.",
    )
    duplicated_raw_ids = raw_inventory.duplicated(
        ["modality", "raw_acquisition_id"]
    ).sum()
    add(
        "raw_acquisition_id_uniqueness",
        "PASS" if duplicated_raw_ids == 0 else "FAIL",
        int(duplicated_raw_ids),
        "Acquisition IDs must be unique within each modality.",
    )
    duplicate_aliases = (
        raw_inventory[
            raw_inventory.duplicated(["modality", "plex", "raw_alias"], keep=False)
        ]
        .groupby(["modality", "plex"])["raw_alias"]
        .agg(lambda x: sorted(set(x)))
        .to_dict()
    )
    add(
        "raw_alias_uniqueness",
        "WARN" if duplicate_aliases else "PASS",
        duplicate_aliases,
        (
            "The protein Samp6-10 workbook repeats aliases tc-885_11 and "
            "tc-885_12; acquisition IDs remain unique."
        ),
    )
    confounded = bool(
        design["condition_perfectly_confounded_with_reporter_tag"].all()
    )
    add(
        "condition_reporter_tag_assignment",
        "WARN" if confounded else "PASS",
        confounded,
        (
            "BL, FL, and GC occupy non-overlapping reporter-tag blocks and the "
            "same map is reused in both plexes (no tag swap)."
        ),
    )
    investigation = INVESTIGATION_FILE.read_text(encoding="utf-8", errors="replace")
    has_tmtpro = "TMTpro reagent" in investigation
    has_tmtpro_mass = "304.207" in investigation
    has_tmtpro_impurity = "isotopic impurities of the different TMTpro" in investigation
    has_itraq = "iTRAQ" in investigation
    add(
        "detailed_protocol_TMTpro_evidence",
        "PASS" if has_tmtpro and has_tmtpro_mass and has_tmtpro_impurity else "FAIL",
        {
            "TMTpro_reagent": has_tmtpro,
            "304.207_Da_tag_mass": has_tmtpro_mass,
            "TMTpro_impurity_correction": has_tmtpro_impurity,
        },
        (
            "Detailed protocol evidence resolves the assay chemistry as TMTpro."
        ),
    )
    add(
        "legacy_investigation_iTRAQ_term",
        "WARN" if has_itraq else "PASS",
        has_itraq,
        (
            "A separate legacy study-level description says iTRAQ; retain this "
            "as a metadata discrepancy rather than merging it with protocol evidence."
        ),
    )
    y_rows = phosphosite_audit[phosphosite_audit["contains_Y65_or_Y68"]]
    y_safe = (
        not y_rows.empty
        and (~y_rows["all_components_strict_canonical"]).all()
        and (~y_rows["Y65_Y68_canonical_SPAK_OSR1_site"]).all()
    )
    add(
        "NCC_Y65_Y68_annotation",
        "PASS" if y_safe else "FAIL",
        int(len(y_rows)),
        "Workbook motifs resolve NCC positions 65/68 as tyrosines, not canonical SPAK/OSR1 sites.",
    )
    mapping_counts = (
        phosphosite_audit["sequence_phosphoform_mapping_status"]
        .value_counts(dropna=False)
        .to_dict()
    )
    all_phosphoforms_mapped = set(mapping_counts) == {"mapped"}
    add(
        "peptide_phosphoform_absolute_site_mapping",
        "PASS" if all_phosphoforms_mapped else "FAIL",
        mapping_counts,
        (
            "Every NCC/SPAK peptide phosphoform must map from # markers and "
            "13-aa motif anchors to absolute residue labels."
        ),
    )
    canonical_residues = phosphosite_audit.loc[
        phosphosite_audit["all_components_strict_canonical"],
        "component_residues",
    ]
    no_canonical_y = not canonical_residues.str.contains("Y", regex=False).any()
    add(
        "strict_canonical_residue_identity",
        "PASS" if no_canonical_y else "FAIL",
        no_canonical_y,
        "Every strict canonical NCC/SPAK feature must be serine/threonine.",
    )
    strict_rows = phosphosite_audit[
        phosphosite_audit["all_components_strict_canonical"]
    ]
    strict_comodified = strict_rows[
        strict_rows[
            "sequence_has_additional_modifications_beyond_feature_components"
        ]
    ]
    add(
        "canonical_feature_sequence_co_modification",
        "WARN" if not strict_comodified.empty else "PASS",
        "|".join(
            (
                strict_comodified["gene_symbol"]
                + ":"
                + strict_comodified["residue_resolved_site_labels"]
                + "->"
                + strict_comodified["sequence_phosphoform_site_labels"]
            ).tolist()
        ),
        (
            "Observed canonical T53 and S383 rollups use peptide sequences with "
            "additional phosphomodifications; describe them as site features, "
            "not isolated single-site occupancy measurements."
        ),
    )
    isolated = phosphosite_audit[
        phosphosite_audit["isolated_canonical_assay_feature"]
    ]
    add(
        "isolated_canonical_assay_coverage",
        "PASS" if not isolated.empty else "WARN",
        int(len(isolated)),
        (
            "No NCC/SPAK workbook row is both canonical-site indexed and an "
            "isolated single-phosphoform feature after sequence provenance."
        ),
    )
    return pd.DataFrame(rows)


def isolated_canonical_assay_features(
    phosphosite_audit: pd.DataFrame,
) -> pd.DataFrame:
    """Return assay rows qualified as isolated canonical-site features.

    The filter requires a strict residue-aware literature match, a single-site
    rollup, and exactly one ``#`` phosphomodification in the reported peptide
    sequence.  It is intentionally empty for the current OSD-462 workbook.
    """
    required = {
        "all_components_strict_canonical",
        "site_feature_kind",
        "n_hash_modifications_in_sequence",
        "isolated_canonical_assay_feature",
    }
    missing = required - set(phosphosite_audit.columns)
    if missing:
        raise ValueError(
            f"phosphosite audit missing qualification columns: {sorted(missing)}"
        )
    return phosphosite_audit[
        phosphosite_audit["isolated_canonical_assay_feature"].astype(bool)
    ].copy()


def source_files() -> list[Path]:
    """Return source files whose hashes belong in the Stage 0 manifest."""
    return [
        PROTEIN_WORKBOOK,
        PHOSPHO_WORKBOOK,
        STUDY_FILE,
        INVESTIGATION_FILE,
        PROTEIN_ASSAY_FILE,
        PHOSPHO_ASSAY_FILE,
        REPO_ROOT
        / "data"
        / "external"
        / "kinase_substrate"
        / "renal_kinase_substrate_core.tsv",
    ]
